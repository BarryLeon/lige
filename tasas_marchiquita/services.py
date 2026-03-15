import io
from decimal import Decimal, InvalidOperation
from datetime import date
from django.db.models import Sum, Min 
import openpyxl
from django.db import transaction
from .models import (
    ArchivoImportacion,
    ResponsablePago,
    Parcela,
    Deuda,
    Cuota,
)


# ──────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────

def _limpiar_str(valor):
    """Convierte a string limpio o vacío."""
    if valor is None:
        return ""
    return str(valor).strip()


def _limpiar_decimal(valor):
    """Convierte a Decimal o None si está vacío."""
    if valor is None or str(valor).strip() == "":
        return None
    try:
        return Decimal(str(valor).strip().replace(",", "."))
    except InvalidOperation:
        return None


def _limpiar_entero(valor):
    """Convierte a int o None si está vacío."""
    if valor is None or str(valor).strip() == "":
        return None
    try:
        return int(valor)
    except (ValueError, TypeError):
        return None


def _es_plan_cuotas(valor):
    """
    Interpreta el campo PLAN DE CUOTAS como booleano.
    - "SI"        → True
    - "NO", vacío → False
    """
    return _limpiar_str(valor).upper() == "SI"


# ──────────────────────────────────────────────────────────────────
# COLUMNAS ESPERADAS EN EL EXCEL
# ──────────────────────────────────────────────────────────────────

COLUMNAS_REQUERIDAS = {
    "NRO_INMUEBLE",
    "RESP_PAGO",
    "SUP_TERRENO",
    "PAR_CATASTRAL",
    "CIRCUNS",
    "FRACCION_NRO",
    "PARCELA_NRO",
    "PARCELA_LET",
    "SUBPARCELA",
    "UNI_FUNCIONAL",
    "VALOR TOTAL_DEUDA",
    "PAGO_TOTAL_DEUDA",
    "PLAN DE CUOTAS",
    "CANTIDAD DE CUOTAS",
    "NUMERO_DE_CUOTA_PAGADA",
    "MONTO_PAGADO",
    "ANTICIPO",
}


def _normalizar(texto):
    """Elimina espacios múltiples internos."""
    return " ".join(str(texto).strip().split())


def _validar_columnas(encabezados):
    """
    Verifica que el archivo tenga todas las columnas requeridas.
    Lanza ValueError con las columnas faltantes si hay alguna ausente.
    Retorna un dict {nombre_columna: índice}.
    """
    encabezados_set = {_normalizar(h) for h in encabezados if h is not None}
    faltantes = COLUMNAS_REQUERIDAS - encabezados_set
    if faltantes:
        raise ValueError(
            f"El archivo no tiene las columnas requeridas: {', '.join(sorted(faltantes))}"
        )
    return {_normalizar(h): idx for idx, h in enumerate(encabezados) if h is not None}

# ──────────────────────────────────────────────────────────────────
# PASO 1: GUARDAR ARCHIVO
#    Solo sube el Excel a la BD. No procesa ninguna fila.
#    Retorna la instancia de ArchivoImportacion creada.
# ──────────────────────────────────────────────────────────────────

def guardar_archivo(archivo_binario: bytes, nombre_archivo: str,
                    periodo: date, usuario=None) -> ArchivoImportacion:
    """
    Guarda el archivo Excel en la BD sin procesarlo.

    Parámetros:
        archivo_binario : contenido del archivo en bytes
        nombre_archivo  : nombre original del archivo
        periodo         : fecha del primer día del mes (ej: date(2024, 3, 1))
        usuario         : instancia de User que realiza la subida (opcional)

    Retorna la instancia de ArchivoImportacion creada.
    """

    # -- Verificar período duplicado (solo si no está revertido) --
    if ArchivoImportacion.objects.filter(periodo=periodo, revertido=False).exists():
        raise ValueError(
            f"Ya existe una importación activa para el período {periodo.strftime('%B %Y')}."
        )

    # -- Abrir el Excel y validar estructura antes de guardar --
    wb = openpyxl.load_workbook(io.BytesIO(archivo_binario), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not filas:
        raise ValueError("El archivo está vacío.")

    _validar_columnas(filas[0])

    # -- Guardar en BD --
    importacion = ArchivoImportacion.objects.create(
        archivo=archivo_binario,
        nombre_archivo=nombre_archivo,
        periodo=periodo,
        importado_por=usuario,
        procesado=False,
    )

    return importacion


# ──────────────────────────────────────────────────────────────────
# PASO 2: PROCESAR ARCHIVO
#    Lee el Excel ya guardado y puebla las tablas de negocio.
# ──────────────────────────────────────────────────────────────────

@transaction.atomic
def procesar_archivo(archivo_id: int) -> dict:
    """
    Procesa un archivo Excel previamente guardado en la BD.
    Primero valida TODAS las filas; si hay algún error no procesa nada
    y retorna el reporte de problemas.

    Retorna un dict con el resumen:
        {
            "archivo_id": int,
            "total_filas": int,
            "creados": int,
            "actualizados": int,
            "errores": [ {"fila": int, "mensaje": str}, ... ],
            "abortado": bool,   # True si se encontraron errores y no se procesó
        }
    """

    # -- Obtener el archivo --
    try:
        importacion = ArchivoImportacion.objects.get(pk=archivo_id)
    except ArchivoImportacion.DoesNotExist:
        raise ValueError(f"No existe un archivo con ID {archivo_id}.")

    if importacion.revertido:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' fue revertido y no puede procesarse."
        )

    if importacion.procesado:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' ya fue procesado anteriormente."
        )

    # -- Leer el Excel desde el binario guardado en BD --
    archivo_binario = bytes(importacion.archivo)
    wb = openpyxl.load_workbook(io.BytesIO(archivo_binario), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    encabezados = filas[0]
    col = _validar_columnas(encabezados)

    # -- Determinar si hay deuda inicial cargada --
    hay_deuda_inicial = ArchivoImportacion.objects.filter(
        procesado=True, revertido=False
    ).exists()

    # ── PASADA 1: validar todas las filas sin escribir nada ───────
    errores = []
    vistos_en_archivo = set()   # para detectar duplicados dentro del mismo archivo
    filas_datos = []            # filas no vacías con su número

    for nro_fila, fila in enumerate(filas[1:], start=2):
        if all(v is None for v in fila):
            continue
        filas_datos.append((nro_fila, fila))

    for nro_fila, fila in filas_datos:
        errores_fila = _validar_fila(
            fila, col, hay_deuda_inicial, vistos_en_archivo
        )
        for msg in errores_fila:
            errores.append({"fila": nro_fila, "mensaje": msg})

    # -- Si hay errores, abortar sin procesar nada --
    if errores:
        return {
            "archivo_id":  importacion.pk,
            "total_filas": len(filas_datos),
            "creados":     0,
            "actualizados": 0,
            "errores":     errores,
            "abortado":    True,
        }

    # ── PASADA 2: procesar todas las filas (sin errores) ─────────
    resumen = {
        "archivo_id":  importacion.pk,
        "total_filas": len(filas_datos),
        "creados":     0,
        "actualizados": 0,
        "errores":     [],
        "abortado":    False,
    }

    for nro_fila, fila in filas_datos:
        try:
            _procesar_fila(fila, col, importacion, resumen)
        except Exception as exc:
            # No debería llegar acá, pero por seguridad
            resumen["errores"].append({"fila": nro_fila, "mensaje": str(exc)})

    # -- Marcar como procesado --
    importacion.total_registros = resumen["total_filas"]
    importacion.procesado = True
    importacion.save(update_fields=["total_registros", "procesado"])

    return resumen


def _extraer_valores_fila(fila, col) -> dict:
    """Extrae y limpia todos los valores de una fila."""
    return {
        "nro_inmueble":    _limpiar_str(fila[col["NRO_INMUEBLE"]]),
        "resp_pago":       _limpiar_str(fila[col["RESP_PAGO"]]),
        "valor_total":     _limpiar_decimal(fila[col["VALOR TOTAL_DEUDA"]]),
        "pago_total":      _limpiar_decimal(fila[col["PAGO_TOTAL_DEUDA"]]),
        "tiene_plan":      _es_plan_cuotas(fila[col["PLAN DE CUOTAS"]]),
        "nro_cuota":       _limpiar_entero(fila[col["NUMERO_DE_CUOTA_PAGADA"]]),
        "monto_pagado":    _limpiar_decimal(fila[col["MONTO_PAGADO"]]),
        "anticipo":        _limpiar_decimal(fila[col["ANTICIPO"]]),
        "circuns":         _limpiar_str(fila[col["CIRCUNS"]]),
        "fraccion_nro":    _limpiar_str(fila[col["FRACCION_NRO"]]),
        "parcela_nro":     _limpiar_str(fila[col["PARCELA_NRO"]]),
        "parcela_let":     _limpiar_str(fila[col["PARCELA_LET"]]),
        "subparcela":      _limpiar_str(fila[col["SUBPARCELA"]]),
    }


def _validar_fila(fila, col, hay_deuda_inicial: bool, vistos_en_archivo: set) -> list:
    """
    Valida una fila contra todas las reglas de negocio.
    No escribe nada en la BD.
    Retorna lista de mensajes de error (vacía si todo está bien).
    """
    errores = []
    v = _extraer_valores_fila(fila, col)

    nro_inmueble = v["nro_inmueble"]
    resp_pago    = v["resp_pago"]
    valor_total  = v["valor_total"]
    tiene_plan   = v["tiene_plan"]
    nro_cuota    = v["nro_cuota"]

    # -- Validaciones básicas --
    if not nro_inmueble:
        errores.append("NRO_INMUEBLE vacío.")
        return errores  # sin identificador no podemos validar más
    if not resp_pago:
        errores.append("RESP_PAGO vacío.")
        return errores
    if valor_total is None:
        errores.append("VALOR TOTAL_DEUDA vacío o inválido.")

    clave = (nro_inmueble, resp_pago)

    # -- Duplicado dentro del mismo archivo --
    if clave in vistos_en_archivo:
        errores.append(
            f"Registro duplicado en el archivo: NRO_INMUEBLE={nro_inmueble} "
            f"RESP_PAGO={resp_pago} aparece más de una vez."
        )
    else:
        vistos_en_archivo.add(clave)

    # -- Validaciones contra la BD --
    parcela_qs = Parcela.objects.filter(
        nro_inmueble=v["nro_inmueble"],
        circuns=v["circuns"],
        fraccion_nro=v["fraccion_nro"],
        parcela_nro=v["parcela_nro"],
        parcela_let=v["parcela_let"],
        subparcela=v["subparcela"],
    )
    parcela = parcela_qs.first()

    if not hay_deuda_inicial:
        # ── Es el archivo inicial ──────────────────────────────
        # No debe existir ya una deuda para esta parcela
        if parcela is not None:
            deuda_existente = Deuda.objects.filter(parcela=parcela).exists()
            if deuda_existente:
                errores.append(
                    f"NRO_INMUEBLE={nro_inmueble} ya tiene deuda registrada. "
                    f"El archivo inicial no puede cargarse dos veces."
                )
    else:
        # ── Es un archivo de pago mensual ──────────────────────
        if parcela is None:
            errores.append(
                f"NRO_INMUEBLE={nro_inmueble} no existe en el padrón. "
                f"Verificá que la deuda inicial haya sido cargada."
            )
        else:
            # Verificar que el responsable coincida
            if parcela.responsable.resp_pago.upper() != resp_pago.upper():
                errores.append(
                    f"NRO_INMUEBLE={nro_inmueble}: el RESP_PAGO '{resp_pago}' "
                    f"no coincide con el registrado '{parcela.responsable.resp_pago}'."
                )

            # Verificar que VALOR_TOTAL_DEUDA coincida con la deuda inicial
            if valor_total is not None:
                deuda_inicial = (
                    Deuda.objects
                    .filter(parcela=parcela)
                    .order_by("importacion__periodo")
                    .first()
                )
                if deuda_inicial and deuda_inicial.valor_total_deuda != valor_total.quantize(Decimal("0.01")):
                    errores.append(
                        f"NRO_INMUEBLE={nro_inmueble}: VALOR_TOTAL_DEUDA={valor_total} "
                        f"no coincide con la deuda inicial registrada "
                        f"({deuda_inicial.valor_total_deuda})."
                    )

            # Verificar cuota duplicada
            if tiene_plan and nro_cuota is not None:
                cuota_existente = Cuota.objects.filter(
                    deuda__parcela=parcela,
                    numero_cuota=nro_cuota,
                ).exists()
                if cuota_existente:
                    errores.append(
                        f"NRO_INMUEBLE={nro_inmueble}: la cuota {nro_cuota} "
                        f"ya fue registrada en un período anterior."
                    )

    return errores


# ──────────────────────────────────────────────────────────────────
# PROCESAMIENTO DE UNA FILA (interno)
# ──────────────────────────────────────────────────────────────────

def _procesar_fila(fila, col, importacion, resumen):
    """
    Procesa una fila del Excel:
      1. get_or_create del ResponsablePago
      2. get_or_create de la Parcela (con detección de cambio de dueño)
      3. create de la Deuda mensual
      4. get_or_create de la Cuota (si tiene plan)
    """

    # ── Extraer valores ───────────────────────────────────────────
    nro_inmueble    = _limpiar_str(fila[col["NRO_INMUEBLE"]])
    resp_pago_val   = _limpiar_str(fila[col["RESP_PAGO"]])
    sup_terreno     = _limpiar_decimal(fila[col["SUP_TERRENO"]])
    par_catastral   = _limpiar_str(fila[col["PAR_CATASTRAL"]])
    circuns         = _limpiar_str(fila[col["CIRCUNS"]])
    fraccion_nro    = _limpiar_str(fila[col["FRACCION_NRO"]])
    parcela_nro     = _limpiar_str(fila[col["PARCELA_NRO"]])
    parcela_let     = _limpiar_str(fila[col["PARCELA_LET"]])
    subparcela      = _limpiar_str(fila[col["SUBPARCELA"]])
    uni_funcional   = _limpiar_str(fila[col["UNI_FUNCIONAL"]])
    valor_total     = _limpiar_decimal(fila[col["VALOR TOTAL_DEUDA"]])
    pago_total      = _limpiar_decimal(fila[col["PAGO_TOTAL_DEUDA"]])
    tiene_plan      = _es_plan_cuotas(fila[col["PLAN DE CUOTAS"]])
    cantidad_cuotas = _limpiar_entero(fila[col["CANTIDAD DE CUOTAS"]])
    nro_cuota       = _limpiar_entero(fila[col["NUMERO_DE_CUOTA_PAGADA"]])
    monto_pagado    = _limpiar_decimal(fila[col["MONTO_PAGADO"]])
    anticipo        = _limpiar_decimal(fila[col["ANTICIPO"]])

    # ── Validaciones mínimas ──────────────────────────────────────
    if not nro_inmueble:
        raise ValueError("NRO_INMUEBLE vacío.")
    if not resp_pago_val:
        raise ValueError("RESP_PAGO vacío.")
    if valor_total is None:
        raise ValueError("VALOR TOTAL_DEUDA vacío o inválido.")

    # ── 1. Responsable de pago ────────────────────────────────────
    responsable, _ = ResponsablePago.objects.get_or_create(
        resp_pago=resp_pago_val
    )

    # ── 2. Parcela ────────────────────────────────────────────────
    parcela, creada = Parcela.objects.get_or_create(
        nro_inmueble=nro_inmueble,
        circuns=circuns,
        fraccion_nro=fraccion_nro,
        parcela_nro=parcela_nro,
        parcela_let=parcela_let,
        subparcela=subparcela,
        defaults={
            "responsable":   responsable,
            "par_catastral": par_catastral,
            "sup_terreno":   sup_terreno,
            "uni_funcional": uni_funcional,
        }
    )

    if creada:
        resumen["creados"] += 1
    else:
        resumen["actualizados"] += 1

        # Detectar cambio de responsable (caso venta de parcela)
        if parcela.responsable_id != responsable.pk:
            HistorialResponsableParcela.objects.create(
                parcela=parcela,
                responsable_anterior=parcela.responsable,
                responsable_nuevo=responsable,
                importacion=importacion,
            )
            parcela.responsable = responsable
            parcela.save(update_fields=["responsable", "actualizado_en"])

    # ── 3. Deuda mensual ──────────────────────────────────────────
    if pago_total is not None and pago_total >= valor_total:
        estado_deuda = Deuda.Estado.CANCELADA
    elif tiene_plan:
        estado_deuda = Deuda.Estado.EN_PLAN
    else:
        estado_deuda = Deuda.Estado.VIGENTE

    deuda = Deuda.objects.create(
        parcela=parcela,
        importacion=importacion,
        valor_total_deuda=valor_total,
        pago_total_deuda=pago_total,
        tiene_plan_cuotas=tiene_plan,
        anticipo=anticipo if tiene_plan else None,
        cantidad_cuotas=cantidad_cuotas if tiene_plan else None,
        estado=estado_deuda,
    )

    # ── 4. Cuota (solo si tiene plan y hay cuota informada) ───────
    if tiene_plan and nro_cuota is not None and monto_pagado is not None:
        Cuota.objects.get_or_create(
            deuda=deuda,
            numero_cuota=nro_cuota,
            defaults={
                "monto_pagado": monto_pagado,
                "estado":       Cuota.Estado.PAGA,
                "importacion":  importacion,
            }
        )

# ──────────────────────────────────────────────────────────────────
# CONSULTAS
# ──────────────────────────────────────────────────────────────────



def _saldo_acumulado(parcela) -> Decimal:
    """
    Calcula el saldo real pendiente de una parcela sumando todo lo
    pagado históricamente: pagos totales + anticipos + cuotas.
    La deuda original se toma del primer archivo (deuda inicial).
    """
    deuda_inicial = (
        Deuda.objects
        .filter(parcela=parcela)
        .order_by("importacion__periodo")
        .first()
    )
    if not deuda_inicial:
        return Decimal("0")

    deuda_original = deuda_inicial.valor_total_deuda

    total_pago = Deuda.objects.filter(parcela=parcela).aggregate(
        t=Sum("pago_total_deuda")
    )["t"] or Decimal("0")

    total_anticipos = Deuda.objects.filter(parcela=parcela).aggregate(
        t=Sum("anticipo")
    )["t"] or Decimal("0")

    total_cuotas = Cuota.objects.filter(
        deuda__parcela=parcela
    ).aggregate(t=Sum("monto_pagado"))["t"] or Decimal("0")

    return deuda_original - total_pago - total_anticipos - total_cuotas


def consulta_por_deudor(resp_pago: str) -> dict:
    """
    Retorna todas las deudas y el historial de pagos de un deudor.
    Incluye el saldo acumulado real por parcela.
    """
    try:
        deudor = ResponsablePago.objects.get(resp_pago__iexact=resp_pago)
    except ResponsablePago.DoesNotExist:
        return {"deudor": None, "parcelas": []}

    parcelas = deudor.parcelas.all()
    resultado = []

    for parcela in parcelas:
        deudas = Deuda.objects.filter(parcela=parcela).order_by("-importacion__periodo")
        cuotas = Cuota.objects.filter(deuda__parcela=parcela).order_by("deuda__importacion__periodo", "numero_cuota")
        resultado.append({
            "parcela":          parcela,
            "deudas":           deudas,
            "cuotas":           cuotas,
            "anticipo_total":   deudas.aggregate(t=Sum("anticipo"))["t"] or Decimal("0"),
            "saldo_acumulado":  _saldo_acumulado(parcela),
        })
    for deuda in deudas:
        cuotas_periodo = Cuota.objects.filter(
            deuda=deuda
        ).aggregate(t=Sum("monto_pagado"))["t"] or Decimal("0")
        deuda.total_pagado_periodo = (deuda.pago_total_deuda or Decimal("0")) + (deuda.anticipo or Decimal("0")) + cuotas_periodo

    return {"deudor": deudor, "parcelas": resultado}


def consulta_por_parcela(nro_inmueble: str) -> dict:
    """
    Retorna el estado actual de una parcela, su deuda más reciente,
    el saldo acumulado real y el detalle del plan de cuotas si tiene.
    """
    try:
        parcela = Parcela.objects.get(nro_inmueble=nro_inmueble)
    except Parcela.DoesNotExist:
        return {"parcela": None, "deuda_actual": None, "cuotas": [], "tiene_plan": False, "saldo_acumulado": Decimal("0")}
    except Parcela.MultipleObjectsReturned:
        parcela = Parcela.objects.filter(nro_inmueble=nro_inmueble).first()

    deuda_actual = (
        Deuda.objects
        .filter(parcela=parcela)
        .order_by("-importacion__periodo")
        .first()
    )

    cuotas = []
    tiene_plan = False

    # Buscar si alguna deuda de esta parcela tiene plan y anticipo
    deuda_con_plan = (
        Deuda.objects
        .filter(parcela=parcela, tiene_plan_cuotas=True)
        .order_by("importacion__periodo")
        .first()
    )

    if deuda_con_plan:
        tiene_plan = True
        anticipo_total = Deuda.objects.filter(parcela=parcela).aggregate(
            t=Sum("anticipo")
        )["t"] or Decimal("0")
        cantidad_cuotas = deuda_con_plan.cantidad_cuotas
        cuotas = Cuota.objects.filter(
            deuda__parcela=parcela
        ).order_by("numero_cuota")
    else:
        anticipo_total  = Decimal("0")
        cantidad_cuotas = None

    return {
        "parcela":          parcela,
        "deuda_actual":     deuda_actual,
        "cuotas":           cuotas,
        "tiene_plan":       tiene_plan,
        "anticipo_total":   anticipo_total,
        "cantidad_cuotas":  cantidad_cuotas,
        "saldo_acumulado":  _saldo_acumulado(parcela),
    }


def totales_generales() -> dict:
    """
    Retorna los totales globales del padrón.
    - Deuda original: valor de la deuda inicial de cada parcela
    - Total cobrado: suma acumulada de pagos + anticipos + cuotas de todos los períodos
    - Con plan / canceladas: basado en la deuda más reciente de cada parcela
    """
    from django.db.models import Max

    if not ArchivoImportacion.objects.filter(procesado=True, revertido=False).exists():
        return {
            "deuda_original":  0,
            "total_cobrado":   0,
            "deuda_pendiente": 0,
            "con_plan_cuotas": 0,
            "canceladas":      0,
        }

    todas_deudas = Deuda.objects.filter(
        importacion__procesado=True,
        importacion__revertido=False,
    )

    # Deuda original: primera deuda de cada parcela
    primeras = (
        todas_deudas
        .values("parcela")
        .annotate(primera_importacion=Min("importacion__periodo"))
    )
    ids_primera_deuda = []
    for item in primeras:
        deuda = todas_deudas.filter(
            parcela_id=item["parcela"],
            importacion__periodo=item["primera_importacion"],
        ).first()
        if deuda:
            ids_primera_deuda.append(deuda.pk)

    deuda_original = Deuda.objects.filter(pk__in=ids_primera_deuda).aggregate(
        t=Sum("valor_total_deuda")
    )["t"] or 0

    # Total cobrado: pagos + anticipos + cuotas de todos los períodos
    total_pago = todas_deudas.aggregate(t=Sum("pago_total_deuda"))["t"] or 0
    total_anticipos = todas_deudas.aggregate(t=Sum("anticipo"))["t"] or 0
    total_cuotas = Cuota.objects.filter(
        importacion__procesado=True,
        importacion__revertido=False,
    ).aggregate(t=Sum("monto_pagado"))["t"] or 0

    total_cobrado   = total_pago + total_anticipos + total_cuotas
    deuda_pendiente = deuda_original - total_cobrado

    # Con plan / canceladas: basado en la deuda más reciente de cada parcela
    ultimas = (
        todas_deudas
        .values("parcela")
        .annotate(ultima_importacion=Max("importacion__periodo"))
    )
    ids_ultima_deuda = []
    for item in ultimas:
        deuda = todas_deudas.filter(
            parcela_id=item["parcela"],
            importacion__periodo=item["ultima_importacion"],
        ).first()
        if deuda:
            ids_ultima_deuda.append(deuda.pk)

    deudas_recientes = Deuda.objects.filter(pk__in=ids_ultima_deuda)
    con_plan_cuotas = deudas_recientes.filter(tiene_plan_cuotas=True).count()
    canceladas      = deudas_recientes.filter(estado=Deuda.Estado.CANCELADA).count()

    return {
        "deuda_original":  deuda_original,
        "total_cobrado":   total_cobrado,
        "deuda_pendiente": deuda_pendiente,
        "con_plan_cuotas": con_plan_cuotas,
        "canceladas":      canceladas,
    }

# ──────────────────────────────────────────────────────────────────
# ELIMINAR ARCHIVO SIN PROCESAR
# ──────────────────────────────────────────────────────────────────

def eliminar_archivo(archivo_id: int) -> str:
    """
    Elimina un archivo que fue subido pero todavía no fue procesado.
    No se puede eliminar un archivo ya procesado.

    Parámetros:
        archivo_id : PK del ArchivoImportacion a eliminar

    Retorna el nombre del archivo eliminado.
    """
    try:
        importacion = ArchivoImportacion.objects.get(pk=archivo_id)
    except ArchivoImportacion.DoesNotExist:
        raise ValueError(f"No existe un archivo con ID {archivo_id}.")

    if importacion.procesado:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' ya fue procesado y no puede eliminarse. "
            f"Si necesitás corregirlo, usá la opción de revertir."
        )

    nombre = importacion.nombre_archivo
    importacion.delete()
    return nombre


# ──────────────────────────────────────────────────────────────────
# REVERTIR PROCESAMIENTO
# ──────────────────────────────────────────────────────────────────

@transaction.atomic
def revertir_procesamiento(archivo_id: int) -> dict:
    """
    Revierte el procesamiento de un archivo, eliminando todas las
    Deudas y Cuotas que generó. Conserva el archivo y lo marca como
    revertido para trazabilidad.
    Parcelas y ResponsablePago NO se eliminan ya que son datos maestros.

    Parámetros:
        archivo_id : PK del ArchivoImportacion a revertir

    Retorna un dict con el resumen:
        {
            "nombre_archivo": str,
            "deudas_eliminadas": int,
            "cuotas_eliminadas": int,
        }
    """
    try:
        importacion = ArchivoImportacion.objects.get(pk=archivo_id)
    except ArchivoImportacion.DoesNotExist:
        raise ValueError(f"No existe un archivo con ID {archivo_id}.")

    if not importacion.procesado:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' no fue procesado, no hay nada que revertir."
        )

    if importacion.revertido:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' ya fue revertido anteriormente."
        )

    # -- Eliminar Cuotas asociadas a las Deudas de este archivo --
    cuotas_eliminadas = Cuota.objects.filter(importacion=importacion).delete()[0]

    # -- Eliminar Deudas de este archivo --
    deudas_eliminadas = Deuda.objects.filter(importacion=importacion).delete()[0]

    # -- Marcar el archivo como revertido --
    importacion.revertido = True
    importacion.procesado = False
    importacion.save(update_fields=["revertido", "procesado"])

    return {
        "nombre_archivo":   importacion.nombre_archivo,
        "deudas_eliminadas": deudas_eliminadas,
        "cuotas_eliminadas": cuotas_eliminadas,
    }

# ──────────────────────────────────────────────────────────────────
# LIQUIDACIÓN
# ──────────────────────────────────────────────────────────────────

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from .models import Liquidacion


@transaction.atomic
def generar_liquidacion(archivo_id: int, usuario=None) -> Liquidacion:
    """
    Calcula la comisión del 25% sobre lo cobrado en el período,
    genera el PDF y guarda la Liquidacion en la BD.

    Parámetros:
        archivo_id : PK del ArchivoImportacion
        usuario    : User que genera la liquidación

    Retorna la instancia de Liquidacion creada.
    """

    try:
        importacion = ArchivoImportacion.objects.get(pk=archivo_id)
    except ArchivoImportacion.DoesNotExist:
        raise ValueError(f"No existe un archivo con ID {archivo_id}.")

    if not importacion.procesado:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' no fue procesado todavía."
        )

    if importacion.revertido:
        raise ValueError(
            f"El archivo '{importacion.nombre_archivo}' fue revertido."
        )

    if hasattr(importacion, 'liquidacion'):
        raise ValueError(
            f"Ya existe una liquidación para el período {importacion.periodo.strftime('%B %Y')}."
        )

    # -- Calcular totales --
    deudas = Deuda.objects.filter(importacion=importacion)

    total_pago_deuda = deudas.aggregate(
        total=Sum("pago_total_deuda")
    )["total"] or Decimal("0")

    total_anticipos = deudas.aggregate(
        total=Sum("anticipo")
    )["total"] or Decimal("0")

    total_cuotas = Cuota.objects.filter(
        importacion=importacion
    ).aggregate(
        total=Sum("monto_pagado")
    )["total"] or Decimal("0")

    total_cobrado = total_pago_deuda + total_anticipos + total_cuotas
    comision      = (total_cobrado * Decimal("0.25")).quantize(Decimal("0.01"))

    # -- Armar detalle por deudor/parcela para el PDF --
    detalle = []
    for deuda in deudas.select_related("parcela__responsable").order_by("parcela__nro_inmueble"):
        cobrado_deuda = Decimal("0")
        if deuda.pago_total_deuda:
            cobrado_deuda += deuda.pago_total_deuda
        if deuda.anticipo:
            cobrado_deuda += deuda.anticipo
        cuotas_deuda = Cuota.objects.filter(
            deuda=deuda, importacion=importacion
        ).aggregate(total=Sum("monto_pagado"))["total"] or Decimal("0")
        cobrado_deuda += cuotas_deuda

        if cobrado_deuda > 0:
            detalle.append({
                "nro_inmueble": deuda.parcela.nro_inmueble,
                "responsable":  deuda.parcela.responsable.resp_pago,
                "cobrado":      cobrado_deuda,
                "comision":     (cobrado_deuda * Decimal("0.25")).quantize(Decimal("0.01")),
            })

    # -- Generar PDF --
    pdf_bytes = _generar_pdf_liquidacion(importacion, total_cobrado, comision,
                                          total_pago_deuda, total_anticipos,
                                          total_cuotas, detalle)

    # -- Guardar en BD --
    liquidacion = Liquidacion.objects.create(
        importacion=importacion,
        total_pago_deuda=total_pago_deuda,
        total_anticipos=total_anticipos,
        total_cuotas=total_cuotas,
        total_cobrado=total_cobrado,
        comision=comision,
        pdf=pdf_bytes,
        generada_por=usuario,
    )

    return liquidacion


def _generar_pdf_liquidacion(importacion, total_cobrado, comision,
                              total_pago_deuda, total_anticipos,
                              total_cuotas, detalle) -> bytes:
    """Genera el PDF de liquidación y retorna los bytes."""

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "titulo",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#2c5364"),
        spaceAfter=6,
    )
    subtitulo_style = ParagraphStyle(
        "subtitulo",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor("#203a43"),
        spaceAfter=4,
    )

    elements = []

    # -- Encabezado --
    elements.append(Paragraph("LIQUIDACIÓN DE COMISIONES", titulo_style))
    elements.append(Paragraph(
        f"Período: {importacion.periodo.strftime('%B %Y').upper()}",
        subtitulo_style
    ))
    elements.append(Paragraph(
        f"Archivo: {importacion.nombre_archivo}",
        subtitulo_style
    ))
    elements.append(Spacer(1, 0.5*cm))

    # -- Resumen de totales --
    elements.append(Paragraph("Resumen", styles["Heading2"]))
    resumen_data = [
        ["Concepto", "Monto"],
        ["Total pagos de deuda", f"$ {total_pago_deuda:,.2f}"],
        ["Total anticipos",      f"$ {total_anticipos:,.2f}"],
        ["Total cuotas pagadas", f"$ {total_cuotas:,.2f}"],
        ["TOTAL COBRADO",        f"$ {total_cobrado:,.2f}"],
        ["COMISIÓN (25%)",       f"$ {comision:,.2f}"],
    ]

    resumen_table = Table(resumen_data, colWidths=[10*cm, 6*cm])
    resumen_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#2c5364")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("ALIGN",        (1, 0), (1, -1),  "RIGHT"),
        ("BACKGROUND",   (0, 4), (-1, 4),  colors.HexColor("#203a43")),
        ("TEXTCOLOR",    (0, 4), (-1, 4),  colors.white),
        ("FONTNAME",     (0, 4), (-1, 4),  "Helvetica-Bold"),
        ("BACKGROUND",   (0, 5), (-1, 5),  colors.HexColor("#e67e22")),
        ("TEXTCOLOR",    (0, 5), (-1, 5),  colors.white),
        ("FONTNAME",     (0, 5), (-1, 5),  "Helvetica-Bold"),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, 3), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.8*cm))

    # -- Detalle por deudor/parcela --
    if detalle:
        elements.append(Paragraph("Detalle por parcela", styles["Heading2"]))
        detalle_data = [["NRO Inmueble", "Responsable", "Cobrado", "Comisión (25%)"]]
        for row in detalle:
            detalle_data.append([
                row["nro_inmueble"],
                row["responsable"],
                f"$ {row['cobrado']:,.2f}",
                f"$ {row['comision']:,.2f}",
            ])

        detalle_table = Table(detalle_data, colWidths=[3*cm, 7*cm, 4*cm, 4*cm])
        detalle_table.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#2c5364")),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ALIGN",          (2, 0), (-1, -1), "RIGHT"),
            ("GRID",           (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("FONTSIZE",       (0, 0), (-1, -1), 9),
        ]))
        elements.append(detalle_table)

    doc.build(elements)
    return buffer.getvalue()


def descargar_pdf_liquidacion(liquidacion_id: int) -> tuple:
    """
    Retorna los bytes del PDF y el nombre de archivo para descarga.

    Retorna: (pdf_bytes, nombre_archivo)
    """
    try:
        liquidacion = Liquidacion.objects.get(pk=liquidacion_id)
    except Liquidacion.DoesNotExist:
        raise ValueError(f"No existe una liquidación con ID {liquidacion_id}.")

    nombre = f"liquidacion_{liquidacion.importacion.periodo.strftime('%Y_%m')}.pdf"
    return bytes(liquidacion.pdf), nombre


def totales_liquidaciones() -> list:
    """
    Retorna el resumen de liquidaciones por mes para la pantalla de consultas.

    Retorna lista ordenada por período desc:
        [
            {
                "periodo":        date,
                "total_cobrado":  Decimal,
                "comision":       Decimal,
            },
            ...
        ]
    """
    liquidaciones = Liquidacion.objects.select_related("importacion").order_by(
        "-importacion__periodo"
    )

    resultado = []
    for liq in liquidaciones:
        resultado.append({
            "periodo":       liq.importacion.periodo,
            "total_cobrado": liq.total_cobrado,
            "comision":      liq.comision,
            "liquidacion_id": liq.pk,
        })

    totales = Liquidacion.objects.aggregate(
        total_cobrado=Sum("total_cobrado"),
        total_comisiones=Sum("comision"),
    )

    return {
        "detalle":          resultado,
        "cantidad":         len(resultado),
        "total_cobrado":    totales["total_cobrado"]   or Decimal("0"),
        "total_comisiones": totales["total_comisiones"] or Decimal("0"),
    }