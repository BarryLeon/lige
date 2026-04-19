"""
exportar.py
Servicios de exportación de carteles a Excel y PDF.
Recibe un queryset ya filtrado desde la vista de informes.
"""

import io
import os
from datetime import datetime
from functools import lru_cache
from math import atan, cos, degrees, floor, log, pi, radians, sinh, tan

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import requests
from PIL import Image as PILImage
from PIL import ImageDraw

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)


TAMANO_TILE_PX = 256
ZOOM_MAPA_CALLES = 16
HEADERS_OSM = {
    "User-Agent": "LIGE/1.0 (+https://municipalidad.local/carteles)",
}


# ── Excel ─────────────────────────────────────────────────────────────────────

def exportar_excel(queryset) -> bytes:
    """
    Genera un archivo Excel con todos los datos del queryset.
    Devuelve los bytes del archivo .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carteles"

    # ── Estilos ───────────────────────────────────────────────────────────────
    color_header = "1E3A5F"
    font_header  = Font(bold=True, color="FFFFFF", size=10)
    fill_header  = PatternFill("solid", fgColor=color_header)
    alin_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alin_izq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fill_par  = PatternFill("solid", fgColor="EEF2F7")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    # ── Encabezados ───────────────────────────────────────────────────────────
    columnas = [
        ("ID",               8),
        ("Kobo ID",          14),
        ("Fecha captura",    16),
        ("Operador",         16),
        ("Tipo cartel",      14),
        ("Estado físico",    14),
        ("Distancia (m)",    13),
        ("Ancho (m)",        11),
        ("Alto (m)",         11),
        ("Superficie (m2)",  14),
        ("Método detección", 16),
        ("Texto OCR",        30),
        ("Observaciones",    30),
        ("Parcela",          20),
        ("Dirección parcela",22),
        ("Prop. terreno",    22),
        ("CUIT prop. terreno",18),
        ("Prop. cartel",     22),
        ("CUIT prop. cartel",18),
        ("Publicidad actual",22),
        ("Lat",              14),
        ("Lon",              14),
        ("Estado proc.",     14),
    ]

    for col_idx, (titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font      = font_header
        celda.fill      = fill_header
        celda.alignment = alin_centro
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Datos ─────────────────────────────────────────────────────────────────
    for fila_idx, c in enumerate(queryset, start=2):
        fill = fill_par if fila_idx % 2 == 0 else fill_impar

        pub = c.publicidad_actual()

        valores = [
            c.id,
            c.kobo_id or "",
            c.fecha.strftime("%d/%m/%Y %H:%M") if c.fecha else "",
            c.operador or "",
            c.get_tipo_cartel_display() if c.tipo_cartel else "",
            c.get_estado_cartel_display() if c.estado_cartel else "",
            c.distancia or "",
            c.ancho_m or "",
            c.alto_m or "",
            c.superficie_m2 or "",
            "",  # método detección — no en modelo, se puede agregar si se guarda
            c.texto_ocr or "",
            c.observaciones or "",
            c.parcela.nomenclatura() if c.parcela else "",
            c.parcela.direccion if c.parcela else "",
            c.parcela.propietario_terreno.nombre_completo() if c.parcela and c.parcela.propietario_terreno else "",
            c.parcela.propietario_terreno.cuit_dni if c.parcela and c.parcela.propietario_terreno else "",
            c.propietario_cartel.nombre_completo() if c.propietario_cartel else "",
            c.propietario_cartel.cuit_dni if c.propietario_cartel else "",
            pub.nombre_empresa() if pub else "",
            c.lat or "",
            c.lon or "",
            c.get_estado_procesamiento_display() if c.estado_procesamiento else "",
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.fill      = fill
            celda.alignment = alin_izq
            celda.font      = Font(size=9)

        ws.row_dimensions[fila_idx].height = 18

    # ── Resumen en hoja 2 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Generado el"
    ws2["B1"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A2"] = "Total registros"
    ws2["B2"] = queryset.count()

    from django.db.models import Sum
    sup = queryset.aggregate(t=Sum("superficie_m2"))["t"]
    ws2["A3"] = "Superficie total (m2)"
    ws2["B3"] = round(sup, 2) if sup else 0

    for celda in ["A1", "A2", "A3"]:
        ws2[celda].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────────

def exportar_pdf(queryset, media_root: str) -> bytes:
    """
    Genera un PDF con tabla de carteles y miniaturas de foto.
    Devuelve los bytes del archivo .pdf.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "titulo",
        parent=styles["Heading1"],
        fontSize=14,
        spaceAfter=6,
        textColor=colors.HexColor("#1E3A5F"),
    )
    estilo_subtitulo = ParagraphStyle(
        "subtitulo",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    estilo_celda = ParagraphStyle(
        "celda",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=10,
    )
    estilo_celda_bold = ParagraphStyle(
        "celdabold",
        parent=estilo_celda,
        fontName="Helvetica-Bold",
    )
    estilo_header = ParagraphStyle(
        "header",
        parent=estilo_celda,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )

    story = []

    # ── Título ────────────────────────────────────────────────────────────────
    story.append(Paragraph("Relevamiento de Carteles Publicitarios", estilo_titulo))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        f"{queryset.count()} registro(s)",
        estilo_subtitulo,
    ))

    # ── Encabezados de tabla ──────────────────────────────────────────────────
    COLOR_HEADER = colors.HexColor("#1E3A5F")
    COLOR_FILA_PAR = colors.HexColor("#EEF2F7")

    encabezados = [
        Paragraph("Foto",          estilo_header),
        Paragraph("#",             estilo_header),
        Paragraph("Tipo / Estado", estilo_header),
        Paragraph("Parcela",       estilo_header),
        Paragraph("Prop. terreno", estilo_header),
        Paragraph("Prop. cartel",  estilo_header),
        Paragraph("Publicidad",    estilo_header),
        Paragraph("Superficie",    estilo_header),
        Paragraph("Texto OCR",     estilo_header),
        Paragraph("Fecha",         estilo_header),
    ]

    # Anchos de columna en cm (total ~25.7cm en A4 landscape con márgenes)
    col_widths = [2.2, 1.0, 2.8, 3.2, 3.2, 3.2, 3.0, 2.0, 4.5, 2.0]

    filas = [encabezados]

    for c in queryset:
        # Miniatura
        img_cell = ""
        if c.foto_anotada and os.path.isfile(c.foto_anotada.path):
            try:
                img_cell = Image(c.foto_anotada.path, width=1.8*cm, height=1.3*cm)
            except Exception:
                img_cell = Paragraph("—", estilo_celda)
        elif c.foto and os.path.isfile(c.foto.path):
            try:
                img_cell = Image(c.foto.path, width=1.8*cm, height=1.3*cm)
            except Exception:
                img_cell = Paragraph("—", estilo_celda)
        else:
            img_cell = Paragraph("Sin foto", estilo_celda)

        pub = c.publicidad_actual()

        tipo_estado = f"{c.get_tipo_cartel_display() or '—'}\n{c.get_estado_cartel_display() or '—'}"
        parcela_txt = ""
        if c.parcela:
            parcela_txt = c.parcela.nomenclatura()
            if c.parcela.direccion:
                parcela_txt += f"\n{c.parcela.direccion}"

        prop_terreno = ""
        if c.parcela and c.parcela.propietario_terreno:
            pt = c.parcela.propietario_terreno
            prop_terreno = f"{pt.nombre_completo()}\n{pt.cuit_dni}"

        prop_cartel = ""
        if c.propietario_cartel:
            pc = c.propietario_cartel
            prop_cartel = f"{pc.nombre_completo()}\n{pc.cuit_dni}"

        superficie = f"{c.superficie_m2:.2f} m2" if c.superficie_m2 else "—"
        fecha_txt  = c.fecha.strftime("%d/%m/%Y") if c.fecha else "—"
        ocr_txt    = (c.texto_ocr or "")[:80]

        fila = [
            img_cell,
            Paragraph(str(c.id), estilo_celda),
            Paragraph(tipo_estado, estilo_celda),
            Paragraph(parcela_txt, estilo_celda),
            Paragraph(prop_terreno, estilo_celda),
            Paragraph(prop_cartel, estilo_celda),
            Paragraph(pub.nombre_empresa() if pub else "—", estilo_celda),
            Paragraph(superficie, estilo_celda_bold),
            Paragraph(ocr_txt, estilo_celda),
            Paragraph(fecha_txt, estilo_celda),
        ]
        filas.append(fila)

    # ── Estilos de tabla ──────────────────────────────────────────────────────
    n_filas = len(filas)
    estilo_tabla = TableStyle([
        # Encabezado
        ("BACKGROUND",  (0, 0), (-1, 0), COLOR_HEADER),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ROWBACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
        # Filas pares
        *[("BACKGROUND", (0, i), (-1, i), COLOR_FILA_PAR)
          for i in range(2, n_filas, 2)],
        # General
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT",   (0, 1), (-1, -1), 38),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
    ])

    tabla = Table(filas, colWidths=[w*cm for w in col_widths], repeatRows=1)
    tabla.setStyle(estilo_tabla)

    story.append(tabla)

    # ── Pie de página con número ──────────────────────────────────────────────
    def pie_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            landscape(A4)[0] - 1.5*cm,
            0.8*cm,
            f"Página {doc.page}"
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=pie_pagina, onLaterPages=pie_pagina)
    return buf.getvalue()


def _crear_miniatura_cartel(cartel, estilo_texto):
    ruta_foto = None
    if cartel.foto_anotada and os.path.isfile(cartel.foto_anotada.path):
        ruta_foto = cartel.foto_anotada.path
    elif cartel.foto and os.path.isfile(cartel.foto.path):
        ruta_foto = cartel.foto.path

    if not ruta_foto:
        return Paragraph("Sin foto disponible", estilo_texto)

    try:
        return Image(ruta_foto, width=8.5 * cm, height=6.0 * cm)
    except Exception:
        return Paragraph("No se pudo cargar la imagen", estilo_texto)


def _parrafo_valor(etiqueta: str, valor: str, estilo):
    return Paragraph(f"<b>{etiqueta}:</b> {valor or '—'}", estilo)


def _lat_lon_a_tile(lat: float, lon: float, zoom: int) -> tuple[float, float]:
    lat_rad = radians(lat)
    n = 2 ** zoom
    x = (lon + 180.0) / 360.0 * n
    y = (1.0 - log(tan(lat_rad) + 1 / cos(lat_rad)) / pi) / 2.0 * n
    return x, y


@lru_cache(maxsize=256)
def _descargar_tile_osm(x: int, y: int, zoom: int) -> PILImage.Image | None:
    max_tile = 2 ** zoom
    if y < 0 or y >= max_tile:
        return None

    x_envuelto = x % max_tile
    url = f"https://tile.openstreetmap.org/{zoom}/{x_envuelto}/{y}.png"

    try:
        respuesta = requests.get(url, headers=HEADERS_OSM, timeout=3)
        respuesta.raise_for_status()
        return PILImage.open(io.BytesIO(respuesta.content)).convert("RGB")
    except Exception:
        return None


def _crear_mapa_ubicacion(lat: float, lon: float) -> io.BytesIO | None:
    ancho_px = 520
    alto_px = 260
    tile_x, tile_y = _lat_lon_a_tile(lat, lon, ZOOM_MAPA_CALLES)
    tile_x_base = floor(tile_x)
    tile_y_base = floor(tile_y)

    mosaico = PILImage.new("RGB", (TAMANO_TILE_PX * 3, TAMANO_TILE_PX * 3), color="white")

    for offset_x in range(-1, 2):
        for offset_y in range(-1, 2):
            tile = _descargar_tile_osm(tile_x_base + offset_x, tile_y_base + offset_y, ZOOM_MAPA_CALLES)
            if tile is None:
                return None
            pos_x = (offset_x + 1) * TAMANO_TILE_PX
            pos_y = (offset_y + 1) * TAMANO_TILE_PX
            mosaico.paste(tile, (pos_x, pos_y))

    centro_x = TAMANO_TILE_PX + (tile_x - tile_x_base) * TAMANO_TILE_PX
    centro_y = TAMANO_TILE_PX + (tile_y - tile_y_base) * TAMANO_TILE_PX
    izquierda = int(round(centro_x - ancho_px / 2))
    superior = int(round(centro_y - alto_px / 2))
    derecha = izquierda + ancho_px
    inferior = superior + alto_px
    recorte = mosaico.crop((izquierda, superior, derecha, inferior))

    draw = ImageDraw.Draw(recorte)
    centro = (ancho_px // 2, alto_px // 2)
    radio = 10
    draw.ellipse(
        (centro[0] - radio, centro[1] - radio, centro[0] + radio, centro[1] + radio),
        fill="#D62828",
        outline="white",
        width=3,
    )

    buffer = io.BytesIO()
    recorte.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def _crear_bloque_ubicacion(cartel, estilo_texto):
    if cartel.lat is None or cartel.lon is None:
        return [Paragraph("Sin coordenadas", estilo_texto)]

    lat = float(cartel.lat)
    lon = float(cartel.lon)
    url_maps = f"https://www.google.com/maps?q={lat},{lon}"
    url_osm = f"https://www.openstreetmap.org/?mlat={lat}&mlon={lon}#map={ZOOM_MAPA_CALLES}/{lat}/{lon}"
    texto_ubicacion = (
        f"<b>Coordenadas:</b> {lat:.6f}, {lon:.6f}<br/>"
        f"<a href=\"{url_maps}\">Ver ubicacion en Google Maps</a><br/>"
        f"<a href=\"{url_osm}\">Ver ubicacion en OpenStreetMap</a><br/>"
        "Mapa base © OpenStreetMap contributors"
    )

    mapa_buffer = _crear_mapa_ubicacion(lat, lon)
    if mapa_buffer is None:
        return [Paragraph(texto_ubicacion, estilo_texto)]

    tabla = Table(
        [[
            Image(mapa_buffer, width=8.8 * cm, height=4.4 * cm),
            Paragraph(texto_ubicacion, estilo_texto),
        ]],
        colWidths=[9.2 * cm, 7.8 * cm],
    )
    tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4DCE5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return [tabla]


def _descripcion_historial(cartel):
    items = []
    for pub in cartel.historial_publicidad.all():
        nombre = pub.nombre_empresa()
        periodo = []
        if pub.fecha_desde:
            periodo.append(f"desde {pub.fecha_desde.strftime('%d/%m/%Y')}")
        if pub.fecha_hasta:
            periodo.append(f"hasta {pub.fecha_hasta.strftime('%d/%m/%Y')}")
        else:
            periodo.append("actual")
        detalle = ", ".join(periodo)
        items.append(f"• {nombre} ({detalle})")
    return "<br/>".join(items) if items else "Sin historial de publicidad"


def exportar_pdf_detallado(queryset, media_root: str) -> bytes:
    """
    Genera un PDF con una ficha detallada por cada cartel del queryset.
    """
    del media_root

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "detalle_titulo",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1E3A5F"),
        spaceAfter=6,
    )
    estilo_subtitulo = ParagraphStyle(
        "detalle_subtitulo",
        parent=styles["Heading2"],
        fontSize=10,
        textColor=colors.HexColor("#36516E"),
        spaceBefore=8,
        spaceAfter=4,
    )
    estilo_texto = ParagraphStyle(
        "detalle_texto",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        spaceAfter=3,
    )
    estilo_ocr = ParagraphStyle(
        "detalle_ocr",
        parent=estilo_texto,
        backColor=colors.HexColor("#F5F7FA"),
        borderPadding=6,
        borderColor=colors.HexColor("#D4DCE5"),
        borderWidth=0.5,
        borderRadius=3,
    )

    story = [
        Paragraph("Fichas detalladas de carteles publicitarios", estilo_titulo),
        Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            f"{queryset.count()} registro(s)",
            estilo_texto,
        ),
        Spacer(1, 0.3 * cm),
    ]

    for indice, cartel in enumerate(queryset, start=1):
        if indice > 1:
            story.append(PageBreak())

        propietario_cartel = cartel.propietario_cartel.nombre_completo() if cartel.propietario_cartel else "Sin propietario asignado"
        propietario_terreno = (
            cartel.parcela.propietario_terreno.nombre_completo()
            if cartel.parcela and cartel.parcela.propietario_terreno
            else "Sin propietario de terreno"
        )
        parcela = cartel.parcela.nomenclatura() if cartel.parcela else "Sin parcela"
        if cartel.parcela and cartel.parcela.direccion:
            parcela = f"{parcela} - {cartel.parcela.direccion}"

        superficie = "—"
        if cartel.superficie_m2:
            superficie = f"{cartel.superficie_m2:.2f} m²"
            if cartel.ancho_m and cartel.alto_m:
                superficie += f" ({cartel.ancho_m:.2f} m x {cartel.alto_m:.2f} m)"

        bloque_datos = [
            _parrafo_valor("Tipo de cartel", cartel.get_tipo_cartel_display() or "—", estilo_texto),
            _parrafo_valor("Estado fisico", cartel.get_estado_cartel_display() or "—", estilo_texto),
            _parrafo_valor("Superficie calculada", superficie, estilo_texto),
            _parrafo_valor("Propietario del cartel", propietario_cartel, estilo_texto),
            _parrafo_valor("Propietario del terreno", propietario_terreno, estilo_texto),
            _parrafo_valor("Parcela", parcela, estilo_texto),
            _parrafo_valor(
                "Fecha de captura",
                cartel.fecha.strftime("%d/%m/%Y %H:%M") if cartel.fecha else "—",
                estilo_texto,
            ),
            _parrafo_valor("Operador", cartel.operador or "—", estilo_texto),
        ]

        tabla_cabecera = Table(
            [[
                _crear_miniatura_cartel(cartel, estilo_texto),
                bloque_datos,
            ]],
            colWidths=[9.0 * cm, 8.4 * cm],
        )
        tabla_cabecera.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4DCE5")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        story.append(Paragraph(
            f"Cartel #{cartel.id} - {cartel.get_tipo_cartel_display() or 'Sin tipo'}",
            estilo_titulo,
        ))
        story.append(tabla_cabecera)
        story.append(Spacer(1, 0.35 * cm))

        story.append(Paragraph("Historial de publicidad", estilo_subtitulo))
        story.append(Paragraph(_descripcion_historial(cartel), estilo_texto))

        story.append(Paragraph("Texto detectado por OCR", estilo_subtitulo))
        story.append(Paragraph(cartel.texto_ocr or "Sin texto OCR detectado", estilo_ocr))

        story.append(Paragraph("Ubicacion", estilo_subtitulo))
        story.extend(_crear_bloque_ubicacion(cartel, estilo_texto))

        if cartel.observaciones:
            story.append(Paragraph("Observaciones", estilo_subtitulo))
            story.append(Paragraph(cartel.observaciones, estilo_texto))

    def pie_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 1.6 * cm, 0.8 * cm, f"Pagina {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=pie_pagina, onLaterPages=pie_pagina)
    return buf.getvalue()
